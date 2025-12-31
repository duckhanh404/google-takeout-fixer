from pathlib import Path
import shutil
from openpyxl import load_workbook


def copy_files_from_excel(
    xlsx_path: str | Path,
    sheet_name: str,
    column_name: str,
    source_dir: str | Path,
    dest_dir: str | Path,
) -> list[str]:
    """
    - Đọc file Excel
    - Lấy danh sách tên file từ 1 cột
    - Gộp với source_dir
    - Copy sang dest_dir

    Trả về: list các file đã copy thành công
    """

    xlsx_path = Path(xlsx_path)
    source_dir = Path(source_dir)
    dest_dir = Path(dest_dir)

    dest_dir.mkdir(parents=True, exist_ok=True)

    # Load workbook
    wb = load_workbook(xlsx_path, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Không tìm thấy sheet: {sheet_name}")

    ws = wb[sheet_name]

    # Tìm index của cột theo tên
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if column_name not in headers:
        raise ValueError(f"Không tìm thấy cột: {column_name}")

    col_idx = headers.index(column_name)

    copied_files = []

    # Duyệt từng dòng dữ liệu
    for row in ws.iter_rows(min_row=2):
        filename = row[col_idx].value

        if not filename:
            continue

        filename = str(filename).strip()
        src_file = source_dir / filename
        dst_file = dest_dir / filename

        if not src_file.exists():
            print(f"❌ Không tìm thấy file: {src_file}")
            continue

        try:
            shutil.copy2(src_file, dst_file)
            copied_files.append(filename)
        except Exception as e:
            print(f"⚠️ Lỗi khi copy {filename}: {e}")

    return copied_files


copied = copy_files_from_excel(
    xlsx_path="main 10122025.xlsx",
    sheet_name="Data",
    column_name="error_media",
    source_dir="E:\Takeout\Google Photos\Anh tu nam 2019",
    dest_dir="E:\loi"
)

print(f"Đã copy {len(copied)} file")