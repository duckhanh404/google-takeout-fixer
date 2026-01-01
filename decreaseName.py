from pathlib import Path
import unicodedata
from typing import List, Optional, Any
from pathlib import Path
from openpyxl import load_workbook
from jsonWork import lists_to_excel

def read_excel_sheet_to_list(
    xlsx_path: str | Path,
    sheet_name: str
) -> List[List[Any]]:
    """
    Đọc toàn bộ dữ liệu của 1 worksheet và trả về dạng list các dòng.

    Mỗi phần tử trong list là 1 list tương ứng với 1 row trong Excel.
    Bao gồm cả header (dòng đầu tiên).
    """

    xlsx_path = Path(xlsx_path)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Không tìm thấy worksheet: {sheet_name}")

    ws = wb[sheet_name]

    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))

    return data


def find_matching_json(
    media_filename: str,
    json_files: List[str]
) -> Optional[str]:
    """
    Tìm file json có tên bắt đầu bằng media name (giảm dần từng ký tự).
    Trả về tên file json nếu tìm thấy, ngược lại trả None.
    """

    # Chuẩn hóa Unicode (rất quan trọng với tiếng Việt)
    media_filename = unicodedata.normalize("NFC", media_filename)
    json_files = [unicodedata.normalize("NFC", f) for f in json_files]

    # Bỏ extension media
    
    base_name = Path(media_filename).stem
    media_ext = Path(media_filename).suffix.lower()
    # Chuẩn bị danh sách json (bỏ .json)
    json_name_map = {
        Path(j).stem: j
        for j in json_files
    }

    # Giảm dần từng ký tự trong media name để tìm
    for i in range(len(base_name), 0, -1):
        prefix = base_name[:i]
        # 
        for json_stem, json_full in json_name_map.items():
            # Nằm ở đầu
            if json_stem.startswith(prefix):
                # cần kiểm tra lại extension trong json nếu cần
                title = unicodedata.normalize("NFC", title)
                title_ext = Path(title).suffix.lower()
                # if media_ext in json_stem.lower():
                return json_full

    return None


if __name__ == "__main__":
    media_rows = read_excel_sheet_to_list("test.xlsx", "media")
    json_rows = read_excel_sheet_to_list("test.xlsx", "json")

    # Bỏ header + lấy cột đầu tiên
    media_files = [row[0] for row in media_rows[1:] if row and row[0]]
    json_files = [row[0] for row in json_rows[1:] if row and row[0]]
    success_count = 0
    success_media = []
    # matched_json = []
    for media in media_files:
        matching_json = find_matching_json(media, json_files)
        if matching_json:
            print(f"Media: {media} -> Matching JSON: {matching_json}")
            success_count += 1
            success_media.append(media)
            success_media.append(matching_json)
            # matched_json.append(matching_json)
        else:
            print(f"Media: {media} -> No matching JSON found.")
    print(f"Loaded {len(media_files)} media files and {len(json_files)} json files.")
    print(f"Total matched: {success_count} out of {len(media_files)}")


    lists_to_excel("decreaseName.xlsx",success_media=success_media)
    # lists_to_excel("decreaseName.xlsx",success_media=success_media, matched_json=matched_json)