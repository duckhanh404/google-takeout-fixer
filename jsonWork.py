import json
import unicodedata
from pathlib import Path 
from datetime import datetime
import re
from openpyxl import Workbook
import subprocess
from test import normalize_existing_path

def extract_title(json_path:Path|str)->str:
    """
    trích xuất title từ file json
    - đầu vào là Path hoặc string đường dẫn đến file json
    - trả về là string đã chuẩn hóa NFC
    """
    json_path = Path(json_path)  # đảm bảo luôn là Path
    with json_path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    title = data.get("title", None)
    # trả về string đã chuẩn hóa NFC
    return unicodedata.normalize("NFC", title)

def extract_title_not(json_path:Path|str)->str:
    """
    trích xuất title từ file json
    - đầu vào là Path hoặc string đường dẫn đến file json
    - trả về là string chưa chuẩn hóa NFC
    """
    json_path = Path(json_path)  # đảm bảo luôn là Path
    with json_path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    title = data.get("title", None)
    # trả về string đã chuẩn hóa NFC
    return title

def extract_time(json_path:Path|str)->str:
    """
    Trích xuất thời gian từ file json
    - đầu vào là Path hoặc string đường dẫn đến file json
    - trả về là timestamp dưới dạng string theo định dạng EXIF: YYYY:MM:DD HH:MM:SS
    Ở đây ưu tiên "photoTakenTime" hơn "creationTime".
    """
    json_path = Path(json_path)
    with json_path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    # ưu tiên đúng thứ tự
    if "photoTakenTime" in data and "timestamp" in data["photoTakenTime"]:
        ts = int(data["photoTakenTime"]["timestamp"])
    elif "creationTime" in data and "timestamp" in data["creationTime"]:
        ts = int(data["creationTime"]["timestamp"])
    else:
        raise ValueError(f"Không có timestamp hợp lệ trong JSON: {json_path}")
    # EXIF format bắt buộc: YYYY:MM:DD HH:MM:SS
    dt = datetime.fromtimestamp(ts)
    return dt.strftime("%Y:%m:%d %H:%M:%S")

def convert_timestamp(timestamp:int|str)->str:
    timestamp = int(timestamp)
    dt = datetime.fromtimestamp(timestamp)
    return dt.strftime("%Y:%m:%d %H:%M:%S")
    

def get_all_json(folder_path:Path|str):
    """
    Lấy tất cả file json trong thư mục
    - đầu vào là Path hoặc string đường dẫn đến thư mục
    - trả về là danh sách tên file json chưa chuẩn hóa
    """
    folder_path = Path(folder_path)
    if not folder_path.is_dir():
        raise ValueError("Đường dẫn không phải thư mục hợp lệ!")

    # return [
    #     unicodedata.normalize("NFC", filename.name)
    #     for filename in folder_path.iterdir()
    #     if filename.is_file() and filename.suffix.lower() == ".json"
    # ]
    return [
        filename.name
        for filename in folder_path.iterdir()
        if filename.is_file() and filename.suffix.lower() == ".json"
    ]

def get_all_media(folder_path:Path|str):
    """
    lấy tất cả file media (không phải json) trong thư mục
    - đầu vào là Path hoặc string đường dẫn đến thư mục
    - trả về là danh sách tên file media chưa chuẩn hóa
    """
    folder_path = Path(folder_path)
    if not folder_path.is_dir():
        raise ValueError("Đường dẫn không hợp lệ!")

    # return [
    #     unicodedata.normalize("NFC", file.name)
    #     for file in folder_path.iterdir()
    #     if file.is_file() and file.suffix.lower() != ".json"
    # ]
    return [
        file.name
        for file in folder_path.iterdir()
        if file.is_file() and file.suffix.lower() != ".json"
    ]

def is_that_cloned(file_name: str) -> bool:
    """
    Kiểm tra tên file có đuôi dạng (số) ngay trước extension hay không.
    Hỗ trợ cả trường hợp có khoảng trắng: "image (1).jpg"
    trả về True nếu đúng là có đuôi (số), ngược lại False.
    Trả về chuỗi "(n)" nếu tên file có dạng (n) ngay trước extension.
    Nếu không có, trả về 0.
    """
    name = Path(file_name).stem

    # Tìm dạng: optional-space + (digits) + end
    match = re.search(r"\s*\(\d+\)$", name)
    if match:
        return match.group().strip()  # return "(2)" (loại bỏ space nếu có)
    return "original"

def lists_to_excel(output_path="output.xlsx", **kwargs):
    """
    Ghi nhiều list vào file Excel.
    Tên mỗi cột chính là tên biến list được truyền vào.
    
    Ví dụ gọi hàm:
    lists_to_excel(output="out.xlsx", list1=a, list2=b, my_data=c)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Lấy danh sách tên cột (keys) và dữ liệu (values)
    headers = list(kwargs.keys())
    columns = list(kwargs.values())

    # Ghi tiêu đề cột
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Tìm số dòng tối đa (list dài nhất)
    max_len = max(len(col) for col in columns)

    # Ghi dữ liệu
    for row_idx in range(max_len):
        for col_idx, col_data in enumerate(columns, start=1):
            if row_idx < len(col_data):
                ws.cell(row=row_idx + 2, column=col_idx, value=col_data[row_idx])

    wb.save(output_path)
    print(f"✔ Đã tạo file Excel: {output_path}")

def check_normalization(s: str):
    """
    Trả về dạng unicode của string: 'NFC', 'NFD', hoặc 'Other'
    """
    if s == unicodedata.normalize("NFC", s):
        return "NFC"
    elif s == unicodedata.normalize("NFD", s):
        return "NFD"
    else:
        return "Other"
    
def get_media_create_timestamp(path: Path | str) -> int | None:
    """
    Lấy MediaCreateDate (sử dụng ModifyDate nếu không có) từ metadata của file và trả về timestamp (int).
    Tùy chọn -charset filename=utf8 được thêm để xử lý tên file có ký tự Unicode.
    """
    
    path_str = str(path)
    
    # 1. Thêm tùy chọn -charset filename=utf8 vào lệnh ExifTool.
    # 2. Sử dụng shell=True, truyền lệnh dưới dạng CHUỖI VÀ BỌC NGOẶC KÉP ("...") 
    #    để đảm bảo tính ổn định cho cả KHOẢNG TRẮNG và UNICODE trên Windows.
    cmd_str = f'exiftool -j -charset filename=utf8 "{path_str}"'
    
    try:
        # CHÚ Ý: Phải đặt shell=True khi truyền chuỗi lệnh
        result = subprocess.run(
            cmd_str, 
            capture_output=True, 
            text=True, 
            shell=True,
            check=True,
            # Chỉ định encoding để xử lý ký tự có dấu trong output và input/output của Shell
            encoding="utf-8" 
        ) 
        
        # Lấy dữ liệu từ output JSON
        data = json.loads(result.stdout)[0]
        
    except subprocess.CalledProcessError as e:
        # Xử lý lỗi khi ExifTool chạy không thành công
        # print(f"❌ ExifTool error (return code {e.returncode}):", e.stderr.strip())
        return None
    except Exception as e:
        # Xử lý các lỗi khác (ví dụ: Lỗi JSON parse)
        # print(f"❌ Lỗi đọc metadata: {e}")
        return None
    
    # --- Phần xử lý thời gian ---
    # Ưu tiên các trường CreateDate, sau đó là ModifyDate
    media_date = data.get("MediaCreateDate") or data.get("CreateDate") or data.get("ModifyDate")

    if not media_date:
        # print("❌ Không tìm thấy thông tin thời gian (MediaCreateDate/CreateDate/ModifyDate) trong metadata.")
        return None

    try:
        # Định dạng thời gian ExifTool trả về thường là "%Y:%m:%d %H:%M:%S"
        # Cần xử lý cả trường hợp có múi giờ hoặc SubSecTime nếu có
        # Ở đây tôi dùng định dạng cơ bản nhất. Nếu lỗi, có thể cần điều chỉnh thêm.
        dt = datetime.strptime(media_date, "%Y:%m:%d %H:%M:%S")
        print(f"✔ Lấy được thời gian từ metadata: {dt} cho file: {path_str}")
        return int(dt.timestamp())
    except Exception as e:
        # print(f"❌ Lỗi parse thời gian: {e}")
        return None

if __name__ == "__main__":
    
    # Đường dẫn có khoảng trắng (raw string)
    test_path = r'/Users/hannada/Downloads/test/số điện thoại.mp4'
    
    # Chạy thử
    a = get_media_create_timestamp(test_path)
    print(a)
