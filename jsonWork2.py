import json
import unicodedata
from pathlib import Path
from datetime import datetime
import re
from openpyxl import Workbook
import subprocess
from exiftool_singleton import ExifToolSingleton

def extract_title(json_path:Path|str)->str:
    """
    trích xuất title từ file json
    - đầu vào là Path hoặc string đường dẫn đến file json
    - trả về là string đã chuẩn hóa NFC
    """
    json_path = Path(json_path)  # đảm bảo luôn là Path
    with json_path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    title = data.get("title", None)
    # trả về string đã chuẩn hóa NFC
    return unicodedata.normalize("NFC", title)

def extract_title_not(json_path:Path|str)->str:
    """
    trích xuất title từ file json
    - đầu vào là Path hoặc string đường dẫn đến file json
    - trả về là string chưa chuẩn hóa NFC
    """
    json_path = Path(json_path)  # đảm bảo luôn là Path
    with json_path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    title = data.get("title", None)
    # trả về string đã chuẩn hóa NFC
    return title

def extract_time(json_path:Path|str)->str:
    """
    Trích xuất thời gian từ file json
    - đầu vào là Path hoặc string đường dẫn đến file json
    - trả về là timestamp dưới dạng string theo định dạng EXIF: YYYY:MM:DD HH:MM:SS
    Ở đây ưu tiên "photoTakenTime" hơn "creationTime".
    """
    json_path = Path(json_path)
    with json_path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    # ưu tiên đúng thứ tự
    if "photoTakenTime" in data and "timestamp" in data["photoTakenTime"]:
        ts = int(data["photoTakenTime"]["timestamp"])
    elif "creationTime" in data and "timestamp" in data["creationTime"]:
        ts = int(data["creationTime"]["timestamp"])
    else:
        raise ValueError(f"Không có timestamp hợp lệ trong JSON: {json_path}")
    # EXIF format bắt buộc: YYYY:MM:DD HH:MM:SS
    dt = datetime.fromtimestamp(ts)
    return dt.strftime("%Y:%m:%d %H:%M:%S")

def convert_timestamp(timestamp:int|str)->str:
    """
    Chuyển timestamp (int hoặc str) thành chuỗi datetime theo định dạng EXIF: YYYY:MM:DD HH:MM:SS
    """
    timestamp = int(timestamp)
    dt = datetime.fromtimestamp(timestamp)
    return dt.strftime("%Y:%m:%d %H:%M:%S")
    
def get_all_json(folder_path:Path|str)->set[str]:
    """
    Lấy tất cả file json trong thư mục
    - đầu vào là Path hoặc string đường dẫn đến thư mục
    - trả về là set tên file json chuẩn hóa
    """
    folder_path = Path(folder_path)
    if not folder_path.is_dir():
        raise ValueError("Đường dẫn không phải thư mục hợp lệ!")

    return {
        unicodedata.normalize("NFC", filename.name)
        for filename in folder_path.iterdir()
        if filename.is_file() and filename.suffix.lower() == ".json"
    }
    # return {
    #     filename.name
    #     for filename in folder_path.iterdir()
    #     if filename.is_file() and filename.suffix.lower() == ".json"
    # }

def get_all_media(folder_path:Path|str)->set[str]:
    """
    lấy tất cả file media (không phải json) trong thư mục
    - đầu vào là Path hoặc string đường dẫn đến thư mục
    - trả về là set tên file media chuẩn hóa
    """
    folder_path = Path(folder_path)
    if not folder_path.is_dir():
        raise ValueError("Đường dẫn không hợp lệ!")

    return {
        unicodedata.normalize("NFC", file.name)
        for file in folder_path.iterdir()
        if file.is_file() and file.suffix.lower() != ".json"
    }
    # return {
    #     file.name
    #     for file in folder_path.iterdir()
    #     if file.is_file() and file.suffix.lower() != ".json"
    # }

def is_that_cloned(file_name: str) -> str:
    """
    Kiểm tra tên file có đuôi dạng (số) ngay trước extension hay không.
    Hỗ trợ cả trường hợp có khoảng trắng: "image (1).jpg"
    trả về True nếu đúng là có đuôi (số), ngược lại False.
    Trả về chuỗi "(n)" nếu tên file có dạng (n) ngay trước extension.
    Nếu không có, trả về 0.
    """
    name = Path(file_name).stem

    # Tìm dạng: optional-space + (digits) + end
    match = re.search(r"\s*\(\d+\)$", name)
    if match:
        return match.group().strip()  # return "(2)" (loại bỏ space nếu có)
    return "original"

def lists_to_excel(output_path="output.xlsx", **kwargs):
    """
    Ghi nhiều list vào file Excel.
    Tên mỗi cột chính là tên biến list được truyền vào.
    
    Ví dụ gọi hàm:
    lists_to_excel(output="out.xlsx", list1=a, list2=b, my_data=c)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Lấy danh sách tên cột (keys) và dữ liệu (values)
    headers = list(kwargs.keys())
    columns = list(kwargs.values())

    # Ghi tiêu đề cột
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Tìm số dòng tối đa (list dài nhất)
    max_len = max(len(col) for col in columns)

    # Ghi dữ liệu
    for row_idx in range(max_len):
        for col_idx, col_data in enumerate(columns, start=1):
            if row_idx < len(col_data):
                ws.cell(row=row_idx + 2, column=col_idx, value=col_data[row_idx])

    wb.save(output_path)
    print(f"✔ Đã tạo file Excel: {output_path}")

def check_normalization(s: str):
    """
    Trả về dạng unicode của string: 'NFC', 'NFD', hoặc 'Other'
    """
    if s == unicodedata.normalize("NFC", s):
        return "NFC"
    elif s == unicodedata.normalize("NFD", s):
        return "NFD"
    else:
        return "Other"

_exiftool = ExifToolSingleton()
def get_media_create_timestamp(path: Path | str) -> int | None:
    path_str = unicodedata.normalize("NFC", str(path))

    try:
        data = _exiftool.execute([
            "-j",
            "-charset", "filename=utf8",
            "-DateTimeOriginal",
            "-CreateDate",
            "-MediaCreateDate",
            "-ModifyDate",
            "-DateCreated",
            path_str
        ])[0]
    except Exception:
        return None

    for key in [
        "DateTimeOriginal",
        "CreateDate",
        "MediaCreateDate",
        "ModifyDate",
        "FileModifyDate"
    ]:
        if key in data:
            dt_str = data[key]
            break
    else:
        return None

    formats = [
        "%Y:%m:%d %H:%M:%S",
        "%Y:%m:%d %H:%M:%S%z",
        "%Y:%m:%d %H:%M:%S.%f",
        "%Y:%m:%d %H:%M:%S.%f%z"
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(dt_str, fmt)
            return int(dt.timestamp())
        except Exception:
            pass

    return None



