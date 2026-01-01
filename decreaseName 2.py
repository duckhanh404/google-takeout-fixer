from pathlib import Path
import unicodedata
from typing import List, Optional, Any
from openpyxl import load_workbook
from jsonWork import lists_to_excel
import json




def read_excel_sheet_to_list(
    xlsx_path: str | Path,
    sheet_name: str
) -> List[List[Any]]:
    """
    Đọc toàn bộ dữ liệu của 1 worksheet và trả về dạng list các dòng.

    Mỗi phần tử trong list là 1 list tương ứng với 1 row trong Excel.
    Bao gồm cả header (dòng đầu tiên).
    """

    xlsx_path = Path(xlsx_path)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Không tìm thấy worksheet: {sheet_name}")

    ws = wb[sheet_name]

    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))

    return data


def find_matching_json(
    media_filename: str,
    json_files: List[str]
) -> Optional[str]:
    """
    Tìm file json có tên bắt đầu bằng media name (giảm dần từng ký tự).
    Chỉ match nếu extension trong field "title" của json
    trùng với extension của media.
    """

    # Chuẩn hóa Unicode (quan trọng với tiếng Việt)
    media_filename = unicodedata.normalize("NFC", media_filename)
    json_files = [unicodedata.normalize("NFC", f) for f in json_files]

    media_path = Path(media_filename)
    base_name = media_path.stem
    media_ext = media_path.suffix.lower()  # .jpg, .mp4, ...

    for i in range(len(base_name), 0, -1):
        prefix = base_name[:i]

        for json_file in json_files:
            json_path = Path(json_file)
            json_stem = json_path.stem

            # Điều kiện 1: tên json bắt đầu bằng prefix
            if not json_stem.startswith(prefix):
                continue

            # Điều kiện 2: đọc field "title" trong json
            try:
                with open(json_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
            except Exception:
                # json lỗi / không đọc được -> bỏ qua
                continue

            title = data.get("title")
            if not isinstance(title, str):
                continue

            title = unicodedata.normalize("NFC", title)
            title_ext = Path(title).suffix.lower()

            # Điều kiện 3: extension phải trùng
            if title_ext == media_ext:
                return json_file

    return None

if __name__ == "__main__":
    media_rows = read_excel_sheet_to_list("test.xlsx", "media")
    json_rows = read_excel_sheet_to_list("test.xlsx", "json")

    # Bỏ header + lấy cột đầu tiên
    media_files = [row[0] for row in media_rows[1:] if row and row[0]]
    json_files = [row[0] for row in json_rows[1:] if row and row[0]]
    success_count = 0
    success_media = []
    # matched_json = []
    for media in media_files:
        matching_json = find_matching_json(media, json_files)
        if matching_json :
            print(f"Media: {media} -> Matching JSON: {matching_json}")
            success_count += 1
            success_media.append(media)
            success_media.append(matching_json)
            # matched_json.append(matching_json)
        else:
            print(f"Media: {media} -> No matching JSON found.")
    print(f"Loaded {len(media_files)} media files and {len(json_files)} json files.")
    print(f"Total matched: {success_count} out of {len(media_files)}")


    lists_to_excel("decreaseName.xlsx",success_media=success_media)
    # lists_to_excel("decreaseName.xlsx",success_media=success_media, matched_json=matched_json)